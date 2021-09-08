from abstract.AbstractParser import AbstractParser
from parser.Exceptions import DataLengthZeroError
from parser.Exceptions import DataLengthUnequalError
from misc.Converters import str_to_integer
import logging
from io import BytesIO
from openpyxl import load_workbook

# NOTE: deprecated as of 2021-09-06, replaced by DailyVaccinationsGithubParser

class DailyVaccinationsParser(AbstractParser):
    def __init__(self, **kwargs):
        self.logger = logging.getLogger(__name__)
        self.logger.debug('__init__() called.')
        self.logger.debug(f'with parameters: {kwargs}')

        for key, value in kwargs.items():
            if key == "source":
                super().__init__(value)
            else:
                setattr(DailyVaccinationsParser, key, value)

    def _parse(self, xmldata):
        self.logger.debug('_parse() called.')
        self.logger.debug(f'with xmldata: {xmldata}')

        with BytesIO(xmldata) as daily_vaccinations:
            wb = load_workbook(daily_vaccinations)
            if wb.sheetnames.count('Impfungen_proTag') != 1:
                raise ValueError('expected excel sheet not found.')

            wb.active = wb['Impfungen_proTag']
            ws = wb.active

            parsed_all_dates = False
            dates = []
            primary_vaccinations = []
            secondary_vaccinations = []

            row_index = 0
            for row in ws:
                if parsed_all_dates is True:
                    break
                # first row is header row and doesn't contain usefull values
                if row_index > 0:
                    col_index = 0
                    row_value = str(row[0].value)
                    # parse rows until end of interesting data
                    if (( row_value is None ) or (row_value == '') or (row_value == 'None')):
                        break
                    for col in row:
                        # the dates column
                        if col_index == 0:
                            day, month, year = None, None, None
                            raw_date_str = str(col.value)
                            raw_date = raw_date_str.split(' ')[0]
                            raw_date_array = raw_date.split('.')
                            if len(raw_date_array) == 3:
                                raw_day = str_to_integer(raw_date_array[0], '+')
                                raw_month = str_to_integer(raw_date_array[1], '+')
                                raw_year = str_to_integer(raw_date_array[2], '+')
                                if (2020 <= raw_year) and (1 <= raw_month <= 12) and (1 <= raw_day <= 31):
                                    if raw_day < 10:
                                        day = f'0{raw_day}'
                                    else:
                                        day = str(raw_day)
                                    if raw_month < 10:
                                        month = f'0{raw_month}'
                                    else:
                                        month = str(raw_month)
                                    year = str(raw_year)
                                    date = f'{year}-{month}-{day}'
                                    dates.append(date)
                                    self.logger.debug(f'appended date: {date}')
                                else:
                                    raise ValueError('day, month or year not in expected range.')
                            elif raw_date_array[0] == 'Gesamt':
                                parsed_all_dates = True
                                break
                            else:
                                raise ValueError('wrong date format.')

                        # the primary vaccinations column
                        elif col_index == 1:
                            raw_p_vacc = str(col.value)
                            p_vacc = str_to_integer(raw_p_vacc, '+')
                            primary_vaccinations.append(p_vacc)
                            self.logger.debug(f'appended p_vacc: {p_vacc}')

                        # the secondary vaccinations column
                        elif col_index == 2:
                            raw_s_vacc = str(col.value)
                            # for dates without secondary vaccinations cell
                            if raw_s_vacc is None or raw_s_vacc == 'None':
                                s_vacc = 0
                                secondary_vaccinations.append(s_vacc)
                                self.logger.debug(f'appended s_vacc: {s_vacc}')
                            else:
                                s_vacc = str_to_integer(raw_s_vacc, '+')
                                secondary_vaccinations.append(s_vacc)
                                self.logger.debug(f'appended s_vacc: {s_vacc}')
                        col_index +=1
                row_index +=1

            ''' check data for consistency, equal amount of dates and cases '''
            if len(dates) != len(primary_vaccinations) != len(secondary_vaccinations):
                raise DataLengthUnequalError('dates, primary_vaccinations and secondary_vaccinations array length not equal.')

            if len(dates) < 1 and len(primary_vaccinations) < 1 and len(secondary_vaccinations):
                raise DataLengthZeroError('dates, primary_vaccinations and secondary_vaccinations array length is zero.')

            dict = { 'dates':dates, 'primary_vaccinations':primary_vaccinations, 'secondary_vaccinations':secondary_vaccinations }
            self.parsed_data = dict
