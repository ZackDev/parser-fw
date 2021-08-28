from abstract.AbstractParser import AbstractParser
from parser.Exceptions import DataLengthZeroError
from parser.Exceptions import DataLengthUnequalError
from parser.Validators import strToInteger
from io import BytesIO
from openpyxl import load_workbook

class WeeklyTestsParser(AbstractParser):
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        super().__init__(source)

    def _parse(self, xmldata):

        self.logger.info('_parse() called.')
        self.logger.debug(f'with xmldata: {xmldata}')

        with BytesIO(xmldata) as weekly_tests:
            wb = load_workbook(weekly_tests)
            if wb.sheetnames.count('1_Testzahlerfassung') != 1:
                raise ValueError('expected excel sheet not found.')

            wb.active = wb['1_Testzahlerfassung']
            ws = wb.active

            calendar_weeks = []
            weekly_tests = []
            parse_error = False

            for x in range(1, 10):
                calendar_weeks.append(f'2020-W0{x}')
                weekly_tests.append(0)

            row_index = 0
            for row in ws:
                if parse_error is True:
                    break
                if row_index == 1:
                    calendar_weeks.append(f'2020-W10')
                    test_count = row[2].value
                    test_count = strToInteger(test_count, '+')
                    weekly_tests.append(test_count)
                    self.logger.debug(f'appended {test_count}')
                elif row_index >= 2:
                    if row[0].value == 'Summe' or row[0].value is None:
                        break
                    col_index = 0
                    for col in row:
                        if col_index == 0:
                            raw_calendar_week = str(col.value)
                            raw_week_array = raw_calendar_week.split('/')
                            raw_week = None
                            raw_year = None
                            if len(raw_week_array) == 2:
                                raw_week = strToInteger(raw_week_array[0], '+')
                                raw_year = strToInteger(raw_week_array[1], '+')
                            else:
                                parse_error = True
                                print('error: parsing raw_week_array.')
                                break
                            if raw_week is not None and raw_week > 0 and raw_week <= 53:
                                if raw_week < 10:
                                    week = f'0{raw_week}'
                                else:
                                    week = f'{raw_week}'
                            else:
                                parse_error = True
                                print('error: parsing raw_week.')
                                break
                            if raw_year is not None and raw_year >= 2020 and raw_year <= 9999:
                                calendar_week = f'{raw_year}-W{week}'
                                calendar_weeks.append(calendar_week)
                                self.logger.debug(f'appended {calendar_week}')
                            else:
                                parse_error = True
                                print('error: parsing raw_year.')
                                break

                        elif col_index == 1:
                            tests = strToInteger(col.value, '+')
                            weekly_tests.append(tests)
                            self.logger.debug(f'appended {tests}')
                        col_index +=1
                row_index +=1

            if parse_error is False:
                ''' data consistency check, length calendar_weeks equals length weekly_tests '''
                if len(calendar_weeks) != len(weekly_tests):
                    raise DataLengthUnequalError(f'calendar_weeks and weekly_tests array length not equal. {len(calendar_weeks)} != {len(weekly_tests)}')

                if len(calendar_weeks) < 1 and len(weekly_tests) < 1:
                    print('no data extracted. ending programm.')
                    raise DataLengthZeroError('calendar_weeks and weekly_tests array length is zero.')

                dict = {'calendar_weeks':calendar_weeks, 'weekly_tests':weekly_tests}
                self.parsed_data = dict

            else:
                print('error parsing weekly tests xslx.')
                raise Exception()
