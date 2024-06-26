from Abstract import AbstractStep, StepError
from misc.Converters import str_to_integer
from io import BytesIO
from openpyxl import load_workbook


class WeeklyTestsParser(AbstractStep):
    def run(self, data):
        with BytesIO(data) as weekly_tests:
            wb = load_workbook(weekly_tests)
            if wb.sheetnames.count('Testzahlen') != 1:
                raise StepError('expected excel sheet not found.')

            wb.active = wb['Testzahlen']
            ws = wb.active

            calendar_weeks = []
            weekly_tests = []
            weekly_tests_positive = []
            parse_error = False

            for x in range(1, 10):
                calendar_weeks.append(f'2020-W0{x}')
                weekly_tests.append(0)
                weekly_tests_positive.append(0)

            row_index = 0
            for row in ws:
                if parse_error is True:
                    break
                if row_index >= 1:
                    col_index = 0
                    for col in row:
                        if col_index == 0:
                            calendar_week = str(col.value)
                            calendar_weeks.append(calendar_week)
                            self.logger.debug(f'appended {calendar_week}')

                        elif col_index == 1:
                            if isinstance(col.value, float):
                                try:
                                    tests = int(col.value)
                                except Exception as e:
                                    raise StepError('unexpected value for tests.') from e
                            else:
                                raise StepError('unexpected type for tests.')
                            if tests is not None:
                                weekly_tests.append(tests)
                                self.logger.debug(f'appended {tests}')
                            else:
                                raise StepError('error reading tests. got "None"')
                        elif col_index == 3:
                            if isinstance(col.value, float):
                                try:
                                    tests_positive = int(col.value)
                                except Exception as e:
                                    raise StepError('unexpected value for tests_positive')
                            else:
                                raise StepError('unexpected type for tests_positive.')
                            if tests_positive is not None:
                                weekly_tests_positive.append(tests_positive)
                            else:
                                raise StepError('error reading tests_positive. got "None"')
                        col_index += 1
                row_index += 1

            if parse_error is False:
                # data consistency check, length calendar_weeks equals length weekly_tests
                if len(calendar_weeks) != len(weekly_tests) != len(weekly_tests_positive) != len(weekly_tests_negative):
                    raise StepError('calendar_weeks, weekly_tests, weekly_tests_positive and weekly_tests_negative array length not equal.')

                if len(calendar_weeks) < 1 and len(weekly_tests) < 1 and len(weekly_tests_positive) < 1 and len(weekly_tests_negative) < 1:
                    raise StepError('calendar_weeks, weekly_tests, weekly_tests_positive and weekly_tests_negative array length is zero.')

                # calculate total tests
                total_tests = []
                for i in range(len(weekly_tests)):
                    total_tests.append(sum(weekly_tests[0:i + 1]))

                # calculate negative tests
                weekly_tests_negative = []
                for i in range(len(weekly_tests)):
                    weekly_tests_negative.append(weekly_tests[i] - weekly_tests_positive[i])

                # build dict
                dict = {
                    "data": []
                }
                for i in range(len(calendar_weeks)):
                    de = {
                        "calendar_week": calendar_weeks[i],
                        "weekly_tests_positive": weekly_tests_positive[i],
                        "weekly_tests_negative": weekly_tests_negative[i],
                        "total_tests": total_tests[i]
                    }
                    dict['data'].append(de)
                return dict

            else:
                raise StepError('error parsing weekly tests xslx.')
